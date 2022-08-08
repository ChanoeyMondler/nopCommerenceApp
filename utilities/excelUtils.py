import openpyxl


def getrowcount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def getcolumncount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_colunm


def readdata(file, sheetname, rowno, columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rowno, column=columno).value


def writedata(file, sheetname, rowno, columno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rowno, column=columno).value = data
    workbook.save(file)

